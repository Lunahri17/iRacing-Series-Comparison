import os
import json
import base64
import hashlib
import requests
import random
from dotenv import load_dotenv
import tkinter as tk
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tkinter import filedialog


# === CARGA VARIABLES DE ENTORNO ===
load_dotenv(dotenv_path=".env")

EMAIL = os.getenv("EMAIL")
PLAIN_PASSWORD = os.getenv("PASSWORD")

SELECTION_FILE = "old_app/selected_series.json"

if not EMAIL or not PLAIN_PASSWORD:
    raise ValueError("Faltan variables de entorno EMAIL o PASSWORD")

# === FUNCIONES AUXILIARES ===

def save_selected_series(selected_list: list):
    """Guarda la selección actual de series en un archivo JSON"""
    with open(SELECTION_FILE, "w", encoding="utf-8") as f:
        json.dump(selected_list, f, indent=2)

def load_selected_series() -> list:
    """Carga las series seleccionadas previamente (si existe el archivo)"""
    if not os.path.exists(SELECTION_FILE):
        return []
    try:
        with open(SELECTION_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def assign_color(track_name: str, used_colors: dict) -> str:
    """Devuelve un color único por circuito"""
    if track_name in used_colors:
        return used_colors[track_name]
    r = lambda: random.randint(180, 240)
    color = f'#{r():02x}{r():02x}{r():02x}'
    used_colors[track_name] = color
    return color

def hash_password(email: str, password: str) -> str:
    """Genera el hash SHA256 en base64 como requiere iRacing"""
    email = email.lower()
    salted = password + email
    hashed = hashlib.sha256(salted.encode("utf-8")).digest()
    return base64.b64encode(hashed).decode("utf-8")

def authenticate(email: str, hashed_password: str) -> requests.Session:
    """Realiza login y devuelve una sesión autenticada"""
    url = "https://members-ng.iracing.com/auth"
    payload = {"email": email, "password": hashed_password}
    session = requests.Session()
    response = session.post(url, json=payload)
    response.raise_for_status()
    return session

def request_to_json_link(json_response: dict) -> dict:
    """Sigue el enlace 'link' del response para obtener el verdadero contenido JSON"""
    if "link" not in json_response:
        raise ValueError("No se encontró la clave 'link' en la respuesta JSON.")
    return requests.get(json_response["link"], timeout=120).json()

# === SCRIPT PRINCIPAL ===

def get_sport_car_series() -> list:
    hashed_pw = hash_password(EMAIL, PLAIN_PASSWORD)
    session = authenticate(EMAIL, hashed_pw)

    # Obtener todas las series/temporadas disponibles
    url = "https://members-ng.iracing.com/data/series/seasons?include_series=true"
    raw_response = session.get(url)
    raw_response.raise_for_status()
    series_data = request_to_json_link(raw_response.json())

    # Filtrar temporadas que usen esas car_class_id
    sport_series = []
    for season in series_data:
        sport_series.append({
            "series_name": season.get("series_name", ""),
            "season_name": season.get("season_name", ""),
            "season_id": season.get("season_id", ""),
            "car_classes": season.get("car_class_ids", []),
            "start_date": season.get("start_date", ""),
            "season_year": season.get("season_year", ""),
            "season_quarter": season.get("season_quarter", ""),
            "schedules": season.get("schedules", [])
        })

    return sport_series

def get_track_schedule_table(series_data: list) -> list:
    """Genera una tabla con (track, fecha, semana) desde schedules de todas las temporadas"""
    table = []

    for season in series_data:
        series_name = season.get("series_name", "")
        season_name = season.get("season_name", "")
        for schedule in season.get("schedules", []):
            track = schedule.get("track", {})
            track_name = track.get("track_name", "Desconocido")
            race_week = schedule.get("race_week_num", "N/A")
            start_date = schedule.get("start_date", "N/A")

            table.append({
                "series_name": series_name,
                "season_name": season_name,
                "track_name": track_name,
                "race_week": race_week,
                "start_date": start_date
            })

    return table

def show_schedule_custom_table(series_data: list):
    """Construye una tabla personalizada con celdas coloreadas por circuito"""

    # Construir estructura pivotada: {fecha: {temporada: circuito}}
    pivot = defaultdict(dict)
    season_names = set()
    all_tracks = set()

    for season in series_data:
        season_name = season.get("season_name", "")
        season_names.add(season_name)
        for sched in season.get("schedules", []):
            date = sched.get("start_date", None)
            if not date:
                continue
            track = sched.get("track", {}).get("track_name", "Desconocido")
            pivot[date][season_name] = track
            all_tracks.add(track)

    sorted_seasons = sorted(season_names)
    sorted_dates = sorted(pivot.keys())
    track_colors = {}
    for track in all_tracks:
        assign_color(track, track_colors)

    # Crear ventana
    root = tk.Tk()
    root.title("Tabla de Circuitos - Celdas Coloreadas por Circuito")

    # Canvas y frame para scroll
    canvas = tk.Canvas(root, borderwidth=0, background="#f0f0f0")
    frame = tk.Frame(canvas, background="#f0f0f0")
    vsb = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
    hsb = tk.Scrollbar(root, orient="horizontal", command=canvas.xview)
    canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

    vsb.pack(side="right", fill="y")
    hsb.pack(side="bottom", fill="x")
    canvas.pack(side="left", fill="both", expand=True)
    canvas.create_window((0, 0), window=frame, anchor="nw")

    def on_frame_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))
    frame.bind("<Configure>", on_frame_configure)

    # === Construir tabla ===

    # Primera fila (encabezado)
    tk.Label(frame, text="Fecha", bg="#d0d0d0", relief="ridge", width=15).grid(row=0, column=0, sticky="nsew")
    for col_idx, season in enumerate(sorted_seasons, start=1):
        tk.Label(frame, text=season, bg="#d0d0d0", relief="ridge", width=25, wraplength=180, justify="center").grid(
            row=0, column=col_idx, sticky="nsew"
        )

    # Filas de contenido
    for row_idx, date in enumerate(sorted_dates, start=1):
        tk.Label(frame, text=date, bg="#eeeeee", relief="ridge", width=15).grid(row=row_idx, column=0, sticky="nsew")
        for col_idx, season in enumerate(sorted_seasons, start=1):
            track = pivot[date].get(season, "")
            bg_color = track_colors.get(track, "#ffffff") if track else "#ffffff"
            label = tk.Label(frame, text=track, bg=bg_color, relief="ridge", width=25, wraplength=180, justify="center")
            label.grid(row=row_idx, column=col_idx, sticky="nsew")

    # Botón exportar a Excel
    def export_action():
        export_schedule_to_excel(pivot, sorted_dates, sorted_seasons, track_colors)

    export_button = tk.Button(root, text="Exportar a Excel", command=export_action, bg="#4CAF50", fg="white")
    export_button.pack(pady=8)

    root.mainloop()

def choose_series_and_show_table(series_data: list):
    """Ventana con checkboxes scrollables y cuadro de búsqueda para seleccionar series"""

    import tkinter.messagebox

    # Mapa completo de todas las temporadas disponibles
    season_map = {
        f"{s['series_name']} - {s['season_name']}": s
        for s in series_data
    }

    all_season_keys = sorted(season_map.keys())
    prev_selection = set(load_selected_series())

    root = tk.Tk()
    root.title("Seleccioná las temporadas a visualizar")

    tk.Label(root, text="Buscar series/temporadas:", font=("Arial", 11)).pack(pady=(10, 0))

    # === Campo de búsqueda ===
    search_var = tk.StringVar()

    search_entry = tk.Entry(root, textvariable=search_var, width=60)
    search_entry.pack(pady=(0, 5), padx=10)

    # === Scrollable frame con Canvas ===
    container = tk.Frame(root)
    container.pack(fill="both", expand=True, padx=10, pady=5)

    canvas = tk.Canvas(container, height=300)
    scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # === Diccionario de checkboxes y variables ===
    var_map = {}

    def render_checkboxes(filter_text=""):
        # Limpiar frame actual
        for widget in scrollable_frame.winfo_children():
            widget.destroy()

        for key in all_season_keys:
            if filter_text.lower() not in key.lower():
                continue
            var = var_map.get(key)
            if var is None:
                var = tk.BooleanVar(value=key in prev_selection)
                var_map[key] = var
            chk = tk.Checkbutton(scrollable_frame, text=key, variable=var, anchor="w", justify="left", wraplength=500)
            chk.pack(anchor="w", padx=5, pady=2)

    def on_search(*args):
        render_checkboxes(search_var.get())

    search_var.trace_add("write", on_search)

    # Inicial
    render_checkboxes()

    # === Botones ===
    button_frame = tk.Frame(root)
    button_frame.pack(pady=10)

    def on_submit():
        selected_keys = [k for k, v in var_map.items() if v.get()]
        if not selected_keys:
            tk.messagebox.showwarning("Advertencia", "Debe seleccionar al menos una serie.")
            return
        selected_seasons = [season_map[k] for k in selected_keys]
        save_selected_series(selected_keys)
        root.destroy()
        show_schedule_custom_table(selected_seasons)

    def select_all():
        for var in var_map.values():
            var.set(True)

    def deselect_all():
        for var in var_map.values():
            var.set(False)
    
    tk.Button(button_frame, text="Seleccionar todo", command=select_all).grid(row=0, column=0, padx=5)
    tk.Button(button_frame, text="Deseleccionar todo", command=deselect_all).grid(row=0, column=1, padx=5)
    tk.Button(button_frame, text="Mostrar tabla", command=on_submit, bg="#007acc", fg="white").grid(row=0, column=2, padx=10)

    root.mainloop()

def export_schedule_to_excel(pivot, sorted_dates, sorted_seasons, track_colors):
    """Exporta la tabla pivotada a un archivo Excel con colores"""

    # Crear libro y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendario iRacing"

    # Encabezados
    ws.cell(row=1, column=1, value="Fecha")
    for col_idx, season in enumerate(sorted_seasons, start=2):
        ws.cell(row=1, column=col_idx, value=season)

    # Celdas con contenido y color
    for row_idx, date in enumerate(sorted_dates, start=2):
        ws.cell(row=row_idx, column=1, value=date)
        for col_idx, season in enumerate(sorted_seasons, start=2):
            track = pivot[date].get(season, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=track)
            if track in track_colors:
                color_hex = track_colors[track].replace("#", "")
                fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                cell.fill = fill

    # Diálogo para guardar
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Guardar como..."
    )

    if file_path:
        wb.save(file_path)


# === EJECUCIÓN ===
def main():
    sport_series = get_sport_car_series()
    if not sport_series:
        print("No se encontraron series con clase Sport Car.")
        return None

    choose_series_and_show_table(sport_series)



if __name__ == "__main__":
    main()
